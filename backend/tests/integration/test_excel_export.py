"""
Integration tests for Excel Export
Test T136: Test openpyxl file generation, verify 8 columns, multiple items per sample
"""
import pytest
from io import BytesIO
from openpyxl import load_workbook
from datetime import datetime

from app.services.excel_service import ExcelExportService
from app.models.check_object import CheckObject, CheckObjectItem


class TestExcelExportIntegration:
    """T136: Integration test for Excel export with openpyxl"""

    def test_generate_excel_file(self, db_session):
        """Test that Excel file is properly generated"""
        # Create test data
        check_object = CheckObject(
            check_no="EXCEL_INT_001",
            sample_name="集成测试样品",
            company_name="测试公司",
            status=1,
            check_result="合格",
            sampling_time=datetime(2024, 1, 15, 10, 30)
        )
        db_session.add(check_object)
        db_session.commit()

        item = CheckObjectItem(
            check_object_id=check_object.id,
            check_item_name="检测项目1",
            check_method="检测方法1",
            check_result="结果值1",
            result_indicator="合格"
        )
        db_session.add(item)
        db_session.commit()

        # Generate Excel
        service = ExcelExportService(db_session)
        excel_bytes = service.export_to_excel([check_object.id])

        # Load and verify
        wb = load_workbook(BytesIO(excel_bytes))
        ws = wb.active

        assert ws is not None
        assert ws.max_row >= 2  # Header + at least 1 data row

    def test_excel_has_8_columns(self, db_session):
        """Test that Excel file has exactly 8 columns"""
        check_object = CheckObject(
            check_no="EXCEL_COL_001",
            sample_name="列测试",
            status=1
        )
        db_session.add(check_object)
        db_session.commit()

        item = CheckObjectItem(
            check_object_id=check_object.id,
            check_item_name="测试项",
            check_method="测试方法"
        )
        db_session.add(item)
        db_session.commit()

        service = ExcelExportService(db_session)
        excel_bytes = service.export_to_excel([check_object.id])

        wb = load_workbook(BytesIO(excel_bytes))
        ws = wb.active

        # Verify 8 columns
        expected_headers = [
            "样品名称", "公司/个体", "检测项目", "检验结果",
            "该项结果", "检测时间", "样品编号", "检测方法"
        ]

        for i, header in enumerate(expected_headers, 1):
            assert ws.cell(row=1, column=i).value == header

        # Column 9 should be empty
        assert ws.cell(row=1, column=9).value is None

    def test_multiple_items_per_sample(self, db_session):
        """Test that multiple check items expand to multiple rows"""
        check_object = CheckObject(
            check_no="EXCEL_MULTI_001",
            sample_name="多项测试",
            company_name="多项公司",
            status=1,
            check_result="合格",
            sampling_time=datetime(2024, 2, 20, 14, 0)
        )
        db_session.add(check_object)
        db_session.commit()

        # Add 5 items
        items = []
        for i in range(5):
            item = CheckObjectItem(
                check_object_id=check_object.id,
                check_item_name=f"检测项目{i+1}",
                check_method=f"方法{i+1}",
                check_result=f"结果{i+1}",
                result_indicator="合格" if i % 2 == 0 else "不合格"
            )
            items.append(item)
            db_session.add(item)
        db_session.commit()

        service = ExcelExportService(db_session)
        excel_bytes = service.export_to_excel([check_object.id])

        wb = load_workbook(BytesIO(excel_bytes))
        ws = wb.active

        # Should have 6 rows: 1 header + 5 data rows
        assert ws.max_row == 6

        # Verify each row has correct data
        for i in range(5):
            row = i + 2  # Data starts at row 2
            assert ws.cell(row=row, column=1).value == "多项测试"  # 样品名称
            assert ws.cell(row=row, column=2).value == "多项公司"  # 公司/个体
            assert ws.cell(row=row, column=3).value == f"检测项目{i+1}"  # 检测项目

    def test_multiple_samples_export(self, db_session):
        """Test exporting multiple samples"""
        samples = []
        for i in range(3):
            check_object = CheckObject(
                check_no=f"EXCEL_BATCH_{i:03d}",
                sample_name=f"批量样品{i+1}",
                company_name=f"公司{i+1}",
                status=1
            )
            db_session.add(check_object)
            db_session.commit()

            item = CheckObjectItem(
                check_object_id=check_object.id,
                check_item_name=f"项目{i+1}",
                check_method=f"方法{i+1}"
            )
            db_session.add(item)
            samples.append(check_object)

        db_session.commit()

        service = ExcelExportService(db_session)
        ids = [s.id for s in samples]
        excel_bytes = service.export_to_excel(ids)

        wb = load_workbook(BytesIO(excel_bytes))
        ws = wb.active

        # Should have 4 rows: 1 header + 3 data rows
        assert ws.max_row == 4

    def test_date_formatting(self, db_session):
        """Test that dates are formatted correctly"""
        check_object = CheckObject(
            check_no="EXCEL_DATE_001",
            sample_name="日期测试",
            status=1,
            sampling_time=datetime(2024, 3, 15, 9, 30, 0)
        )
        db_session.add(check_object)
        db_session.commit()

        item = CheckObjectItem(
            check_object_id=check_object.id,
            check_item_name="测试项"
        )
        db_session.add(item)
        db_session.commit()

        service = ExcelExportService(db_session)
        excel_bytes = service.export_to_excel([check_object.id])

        wb = load_workbook(BytesIO(excel_bytes))
        ws = wb.active

        # Column 6 is 检测时间
        date_value = ws.cell(row=2, column=6).value
        assert date_value is not None
        # Should contain date info
        assert "2024" in str(date_value) or "03" in str(date_value) or "15" in str(date_value)

    def test_empty_check_items(self, db_session):
        """Test sample with no check items"""
        check_object = CheckObject(
            check_no="EXCEL_EMPTY_001",
            sample_name="空项目测试",
            status=1
        )
        db_session.add(check_object)
        db_session.commit()

        service = ExcelExportService(db_session)
        excel_bytes = service.export_to_excel([check_object.id])

        wb = load_workbook(BytesIO(excel_bytes))
        ws = wb.active

        # Should still have data (maybe 1 row with empty item fields)
        assert ws.max_row >= 1

    def test_chinese_characters(self, db_session):
        """Test that Chinese characters are properly encoded"""
        check_object = CheckObject(
            check_no="EXCEL_ZH_001",
            sample_name="中文样品名称测试",
            company_name="中文公司名称有限公司",
            status=1,
            check_result="合格"
        )
        db_session.add(check_object)
        db_session.commit()

        item = CheckObjectItem(
            check_object_id=check_object.id,
            check_item_name="中文检测项目",
            check_method="中文检测方法",
            check_result="中文结果值",
            result_indicator="合格"
        )
        db_session.add(item)
        db_session.commit()

        service = ExcelExportService(db_session)
        excel_bytes = service.export_to_excel([check_object.id])

        wb = load_workbook(BytesIO(excel_bytes))
        ws = wb.active

        # Verify Chinese content
        assert ws.cell(row=2, column=1).value == "中文样品名称测试"
        assert ws.cell(row=2, column=2).value == "中文公司名称有限公司"
        assert ws.cell(row=2, column=3).value == "中文检测项目"

    def test_special_characters_handling(self, db_session):
        """Test handling of special characters"""
        check_object = CheckObject(
            check_no="EXCEL_SPEC_001",
            sample_name="样品<>&\"'",
            company_name="公司\n换行\t制表",
            status=1
        )
        db_session.add(check_object)
        db_session.commit()

        item = CheckObjectItem(
            check_object_id=check_object.id,
            check_item_name="项目"
        )
        db_session.add(item)
        db_session.commit()

        service = ExcelExportService(db_session)
        excel_bytes = service.export_to_excel([check_object.id])

        # Should not raise exception
        wb = load_workbook(BytesIO(excel_bytes))
        ws = wb.active
        assert ws.max_row >= 2

    def test_row_count_calculation(self, db_session):
        """Test that row count is calculated correctly for limit validation"""
        # Create 10 samples with 5 items each = 50 rows
        for i in range(10):
            check_object = CheckObject(
                check_no=f"EXCEL_COUNT_{i:03d}",
                sample_name=f"计数样品{i}",
                status=1
            )
            db_session.add(check_object)
            db_session.commit()

            for j in range(5):
                item = CheckObjectItem(
                    check_object_id=check_object.id,
                    check_item_name=f"项目{j}"
                )
                db_session.add(item)
        db_session.commit()

        service = ExcelExportService(db_session)

        # Get all IDs
        all_objects = db_session.query(CheckObject).filter(
            CheckObject.check_no.like("EXCEL_COUNT_%")
        ).all()
        ids = [obj.id for obj in all_objects]

        row_count = service.calculate_row_count(ids)
        assert row_count == 50  # 10 samples * 5 items
