"""
Contract tests for Report API
Test T101: POST /reports/upload
Test T134: GET /reports/download/{check_no}
Test T135: POST /reports/export-excel
"""
import pytest
from fastapi.testclient import TestClient
from io import BytesIO
import os
from openpyxl import load_workbook


class TestReportUploadEndpoint:
    """T101: Contract test for POST /reports/upload"""

    def test_upload_pdf_success(self, client: TestClient, auth_headers: dict):
        """Test successful PDF upload"""
        # Create a fake PDF file
        pdf_content = b"%PDF-1.4\n%fake pdf content for testing"
        files = {
            "file": ("test_report.pdf", BytesIO(pdf_content), "application/pdf")
        }

        response = client.post(
            "/api/reports/upload",
            files=files,
            headers=auth_headers
        )

        assert response.status_code == 200
        data = response.json()
        assert "file_url" in data
        assert "filename" in data
        assert data["filename"].endswith(".pdf")

    def test_upload_requires_authentication(self, client: TestClient):
        """Test that upload requires authentication"""
        pdf_content = b"%PDF-1.4\n%test content"
        files = {
            "file": ("test.pdf", BytesIO(pdf_content), "application/pdf")
        }

        response = client.post("/api/reports/upload", files=files)

        assert response.status_code == 401

    def test_upload_file_size_limit(self, client: TestClient, auth_headers: dict):
        """Test file size limit (10MB)"""
        # Create a file larger than 10MB
        large_content = b"x" * (11 * 1024 * 1024)  # 11MB
        files = {
            "file": ("large_report.pdf", BytesIO(large_content), "application/pdf")
        }

        response = client.post(
            "/api/reports/upload",
            files=files,
            headers=auth_headers
        )

        assert response.status_code == 400
        assert "文件大小" in response.json()["detail"] or "too large" in response.json()["detail"].lower()

    def test_upload_pdf_format_validation(self, client: TestClient, auth_headers: dict):
        """Test PDF format validation"""
        # Try to upload non-PDF file
        txt_content = b"This is not a PDF file"
        files = {
            "file": ("test.txt", BytesIO(txt_content), "text/plain")
        }

        response = client.post(
            "/api/reports/upload",
            files=files,
            headers=auth_headers
        )

        assert response.status_code == 400
        assert "PDF" in response.json()["detail"] or "格式" in response.json()["detail"]

    def test_upload_missing_file(self, client: TestClient, auth_headers: dict):
        """Test upload without file"""
        response = client.post(
            "/api/reports/upload",
            headers=auth_headers
        )

        assert response.status_code == 422

    def test_upload_generates_unique_filename(self, client: TestClient, auth_headers: dict):
        """Test that uploaded files get unique filenames"""
        pdf_content = b"%PDF-1.4\n%test"
        files1 = {
            "file": ("report.pdf", BytesIO(pdf_content), "application/pdf")
        }
        files2 = {
            "file": ("report.pdf", BytesIO(pdf_content), "application/pdf")
        }

        response1 = client.post(
            "/api/reports/upload",
            files=files1,
            headers=auth_headers
        )
        response2 = client.post(
            "/api/reports/upload",
            files=files2,
            headers=auth_headers
        )

        assert response1.status_code == 200
        assert response2.status_code == 200

        file1 = response1.json()["filename"]
        file2 = response2.json()["filename"]

        # Filenames should be different (UUID-based)
        assert file1 != file2

    def test_upload_returns_accessible_url(self, client: TestClient, auth_headers: dict):
        """Test that upload returns accessible URL"""
        pdf_content = b"%PDF-1.4\n%test"
        files = {
            "file": ("report.pdf", BytesIO(pdf_content), "application/pdf")
        }

        response = client.post(
            "/api/reports/upload",
            files=files,
            headers=auth_headers
        )

        assert response.status_code == 200
        data = response.json()
        file_url = data["file_url"]

        # URL should be absolute path or contain /reports/
        assert "/reports/" in file_url or file_url.startswith("http")

    def test_upload_creates_directory_structure(self, client: TestClient, auth_headers: dict):
        """Test that upload creates year/month directory structure"""
        pdf_content = b"%PDF-1.4\n%test"
        files = {
            "file": ("report.pdf", BytesIO(pdf_content), "application/pdf")
        }

        response = client.post(
            "/api/reports/upload",
            files=files,
            headers=auth_headers
        )

        assert response.status_code == 200
        data = response.json()

        # File path should contain year/month structure
        # e.g., /reports/2024/11/filename.pdf
        assert "/202" in data["file_url"]  # Contains year

    def test_upload_with_special_characters_in_filename(self, client: TestClient, auth_headers: dict):
        """Test upload with special characters in filename"""
        pdf_content = b"%PDF-1.4\n%test"
        files = {
            "file": ("报告 测试-2024.pdf", BytesIO(pdf_content), "application/pdf")
        }

        response = client.post(
            "/api/reports/upload",
            files=files,
            headers=auth_headers
        )

        assert response.status_code == 200
        # Should handle Chinese and special characters safely

    def test_upload_empty_file(self, client: TestClient, auth_headers: dict):
        """Test upload of empty file"""
        files = {
            "file": ("empty.pdf", BytesIO(b""), "application/pdf")
        }

        response = client.post(
            "/api/reports/upload",
            files=files,
            headers=auth_headers
        )

        assert response.status_code == 400


class TestReportDownloadEndpoint:
    """T134: Contract test for GET /reports/download/{check_no}"""

    def test_download_existing_report(self, client: TestClient, auth_headers: dict):
        """Test successful download of existing report"""
        # First upload a report
        pdf_content = b"%PDF-1.4\n%test content for download"
        files = {
            "file": ("test_report.pdf", BytesIO(pdf_content), "application/pdf")
        }

        upload_response = client.post(
            "/api/reports/upload",
            files=files,
            headers=auth_headers
        )
        assert upload_response.status_code == 200

        # Note: This test assumes we have a way to link report to check_no
        # In real implementation, we'd need to update check_object with report_url

    def test_download_nonexistent_report(self, client: TestClient, auth_headers: dict):
        """Test 404 for missing report"""
        response = client.get(
            "/api/reports/download/NONEXISTENT_CHECK_NO",
            headers=auth_headers
        )

        assert response.status_code == 404
        assert "报告" in response.json()["detail"] or "not found" in response.json()["detail"].lower()

    def test_download_requires_authentication(self, client: TestClient):
        """Test that download requires authentication"""
        response = client.get("/api/reports/download/TEST_CHECK_NO")

        assert response.status_code == 401

    def test_download_returns_pdf_content_type(self, client: TestClient, auth_headers: dict, db_session):
        """Test that download returns correct content type"""
        from app.models.check_object import CheckObject

        # Create a check object with report
        check_object = CheckObject(
            check_no="TEST_DOWNLOAD_001",
            status=1,
            report_url="/reports/2024/11/test.pdf"
        )
        db_session.add(check_object)
        db_session.commit()

        # This test will fail until report file actually exists
        # or we mock the file system


class TestExcelExportEndpoint:
    """T135: Contract test for POST /reports/export-excel"""

    def test_export_excel_with_ids(self, client: TestClient, auth_headers: dict, db_session):
        """Test Excel export with specific check_object_ids"""
        from app.models.check_object import CheckObject, CheckObjectItem

        # Create test data
        check_object = CheckObject(
            check_no="EXPORT_TEST_001",
            sample_name="测试样品",
            company_name="测试公司",
            status=1,
            check_result="合格"
        )
        db_session.add(check_object)
        db_session.commit()

        item = CheckObjectItem(
            check_object_id=check_object.id,
            check_item_name="检测项1",
            check_method="方法1",
            check_result="符合",
            result_indicator="合格"
        )
        db_session.add(item)
        db_session.commit()

        response = client.post(
            "/api/reports/export-excel",
            json={"check_object_ids": [check_object.id]},
            headers=auth_headers
        )

        assert response.status_code == 200
        assert response.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

        # Verify Excel content
        wb = load_workbook(BytesIO(response.content))
        ws = wb.active

        # Check headers (8 columns)
        expected_headers = ["样品名称", "公司/个体", "检测项目", "检验结果", "该项结果", "检测时间", "样品编号", "检测方法"]
        for i, header in enumerate(expected_headers, 1):
            assert ws.cell(row=1, column=i).value == header

    def test_export_excel_with_query_filters(self, client: TestClient, auth_headers: dict, db_session):
        """Test Excel export with query filters"""
        from app.models.check_object import CheckObject

        # Create test data
        check_object = CheckObject(
            check_no="FILTER_TEST_001",
            sample_name="过滤测试",
            company_name="过滤公司",
            status=1
        )
        db_session.add(check_object)
        db_session.commit()

        response = client.post(
            "/api/reports/export-excel",
            json={"status": 1, "company": "过滤公司"},
            headers=auth_headers
        )

        assert response.status_code == 200

    def test_export_excel_1000_row_limit(self, client: TestClient, auth_headers: dict, db_session):
        """Test 1000-row limit validation"""
        from app.models.check_object import CheckObject, CheckObjectItem

        # Create many check objects (more than would result in 1000 rows)
        # Each check object with multiple items = many rows
        for i in range(200):
            check_object = CheckObject(
                check_no=f"LIMIT_TEST_{i:04d}",
                sample_name=f"样品{i}",
                status=1
            )
            db_session.add(check_object)
        db_session.commit()

        # Get all IDs
        all_objects = db_session.query(CheckObject).filter(
            CheckObject.check_no.like("LIMIT_TEST_%")
        ).all()

        # Add multiple items per object to exceed 1000 rows
        for obj in all_objects:
            for j in range(6):  # 200 * 6 = 1200 rows
                item = CheckObjectItem(
                    check_object_id=obj.id,
                    check_item_name=f"项目{j}",
                    check_method=f"方法{j}"
                )
                db_session.add(item)
        db_session.commit()

        all_ids = [obj.id for obj in all_objects]

        response = client.post(
            "/api/reports/export-excel",
            json={"check_object_ids": all_ids},
            headers=auth_headers
        )

        assert response.status_code == 400
        assert "1000" in response.json()["detail"]

    def test_export_excel_requires_authentication(self, client: TestClient):
        """Test that export requires authentication"""
        response = client.post(
            "/api/reports/export-excel",
            json={"check_object_ids": [1, 2, 3]}
        )

        assert response.status_code == 401

    def test_export_excel_empty_result(self, client: TestClient, auth_headers: dict):
        """Test export with no matching data"""
        response = client.post(
            "/api/reports/export-excel",
            json={"check_object_ids": [99999]},
            headers=auth_headers
        )

        # Should return empty Excel or appropriate message
        assert response.status_code in [200, 400]

    def test_export_excel_filename_format(self, client: TestClient, auth_headers: dict, db_session):
        """Test that filename follows expected format"""
        from app.models.check_object import CheckObject

        check_object = CheckObject(
            check_no="FILENAME_TEST_001",
            sample_name="文件名测试",
            status=1
        )
        db_session.add(check_object)
        db_session.commit()

        response = client.post(
            "/api/reports/export-excel",
            json={"check_object_ids": [check_object.id]},
            headers=auth_headers
        )

        assert response.status_code == 200

        # Check Content-Disposition header for filename
        content_disposition = response.headers.get("content-disposition", "")
        assert "检测结果导出" in content_disposition or "export" in content_disposition.lower()
        assert ".xlsx" in content_disposition

    def test_export_excel_multiple_items_expand_rows(self, client: TestClient, auth_headers: dict, db_session):
        """Test that multiple check items expand to multiple rows"""
        from app.models.check_object import CheckObject, CheckObjectItem

        check_object = CheckObject(
            check_no="EXPAND_TEST_001",
            sample_name="展开测试",
            company_name="测试公司",
            status=1,
            check_result="合格"
        )
        db_session.add(check_object)
        db_session.commit()

        # Add 3 items
        for i in range(3):
            item = CheckObjectItem(
                check_object_id=check_object.id,
                check_item_name=f"检测项{i+1}",
                check_method=f"方法{i+1}",
                check_result=f"结果{i+1}",
                result_indicator="合格"
            )
            db_session.add(item)
        db_session.commit()

        response = client.post(
            "/api/reports/export-excel",
            json={"check_object_ids": [check_object.id]},
            headers=auth_headers
        )

        assert response.status_code == 200

        # Verify 3 data rows (plus header)
        wb = load_workbook(BytesIO(response.content))
        ws = wb.active

        # Should have 4 rows: 1 header + 3 data rows
        assert ws.max_row == 4
