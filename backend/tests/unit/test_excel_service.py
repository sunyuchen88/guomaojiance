"""
Unit tests for Excel Export Service
Test T137: Test row expansion logic, column formatting
"""
import pytest
from unittest.mock import Mock, MagicMock
from datetime import datetime
from io import BytesIO
from openpyxl import load_workbook

from app.services.excel_service import ExcelExportService
from app.models.check_object import CheckObject, CheckObjectItem


class TestExcelExportService:
    """T137: Unit test for Excel export service"""

    def test_expand_check_items_to_rows(self, db_session):
        """Test row expansion logic for check items"""
        # Create mock data
        check_object = CheckObject(
            check_no="ROW_EXPAND_001",
            sample_name="行展开测试",
            company_name="测试公司",
            status=1,
            check_result="合格",
            sampling_time=datetime(2024, 1, 1, 12, 0)
        )
        db_session.add(check_object)
        db_session.commit()

        # Add 3 items
        for i in range(3):
            item = CheckObjectItem(
                check_object_id=check_object.id,
                check_item_name=f"项目{i}",
                check_method=f"方法{i}",
                check_result=f"结果{i}",
                result_indicator="合格"
            )
            db_session.add(item)
        db_session.commit()

        service = ExcelExportService(db_session)
        rows = service.expand_check_items_to_rows([check_object])

        # Should have 3 rows (one per item)
        assert len(rows) == 3

        # All rows should have same sample info
        for row in rows:
            assert row["sample_name"] == "行展开测试"
            assert row["company_name"] == "测试公司"
            assert row["check_result"] == "合格"
            assert row["check_no"] == "ROW_EXPAND_001"

        # Each row should have different item info
        assert rows[0]["check_item_name"] == "项目0"
        assert rows[1]["check_item_name"] == "项目1"
        assert rows[2]["check_item_name"] == "项目2"

    def test_sample_without_items_creates_one_row(self, db_session):
        """Test that sample without check items still creates one row"""
        check_object = CheckObject(
            check_no="NO_ITEMS_001",
            sample_name="无项目样品",
            status=1
        )
        db_session.add(check_object)
        db_session.commit()

        service = ExcelExportService(db_session)
        rows = service.expand_check_items_to_rows([check_object])

        # Should have 1 row with empty item fields
        assert len(rows) == 1
        assert rows[0]["sample_name"] == "无项目样品"
        assert rows[0]["check_item_name"] in [None, ""]

    def test_column_formatting(self, db_session):
        """Test that columns are formatted correctly"""
        check_object = CheckObject(
            check_no="FORMAT_001",
            sample_name="格式测试",
            company_name="格式公司",
            status=1,
            check_result="合格",
            sampling_time=datetime(2024, 5, 10, 14, 30)
        )
        db_session.add(check_object)
        db_session.commit()

        item = CheckObjectItem(
            check_object_id=check_object.id,
            check_item_name="格式项目",
            check_method="格式方法",
            check_result="格式结果",
            result_indicator="合格"
        )
        db_session.add(item)
        db_session.commit()

        service = ExcelExportService(db_session)
        excel_bytes = service.export_to_excel([check_object.id])

        wb = load_workbook(BytesIO(excel_bytes))
        ws = wb.active

        # Verify data row
        assert ws.cell(row=2, column=1).value == "格式测试"  # 样品名称
        assert ws.cell(row=2, column=2).value == "格式公司"  # 公司/个体
        assert ws.cell(row=2, column=3).value == "格式项目"  # 检测项目
        assert ws.cell(row=2, column=4).value == "合格"  # 检验结果
        assert ws.cell(row=2, column=5).value == "格式结果"  # 该项结果
        # Column 6 is 检测时间
        assert ws.cell(row=2, column=7).value == "FORMAT_001"  # 样品编号
        assert ws.cell(row=2, column=8).value == "格式方法"  # 检测方法

    def test_date_column_formatting(self, db_session):
        """Test date column formatting"""
        check_object = CheckObject(
            check_no="DATE_FORMAT_001",
            sample_name="日期格式测试",
            status=1,
            sampling_time=datetime(2024, 12, 25, 15, 45, 30)
        )
        db_session.add(check_object)
        db_session.commit()

        item = CheckObjectItem(
            check_object_id=check_object.id,
            check_item_name="项目"
        )
        db_session.add(item)
        db_session.commit()

        service = ExcelExportService(db_session)
        rows = service.expand_check_items_to_rows([check_object])

        # Date should be formatted
        date_value = rows[0]["sampling_time"]
        assert date_value is not None

    def test_null_value_handling(self, db_session):
        """Test that null values are handled properly"""
        check_object = CheckObject(
            check_no="NULL_TEST_001",
            sample_name=None,  # Null sample name
            company_name=None,
            status=0,
            check_result=None,
            sampling_time=None
        )
        db_session.add(check_object)
        db_session.commit()

        item = CheckObjectItem(
            check_object_id=check_object.id,
            check_item_name=None,
            check_method=None,
            check_result=None,
            result_indicator=None
        )
        db_session.add(item)
        db_session.commit()

        service = ExcelExportService(db_session)
        excel_bytes = service.export_to_excel([check_object.id])

        # Should not raise exception
        wb = load_workbook(BytesIO(excel_bytes))
        ws = wb.active
        assert ws.max_row >= 2

    def test_calculate_row_count_single_sample(self, db_session):
        """Test row count calculation for single sample"""
        check_object = CheckObject(
            check_no="COUNT_SINGLE_001",
            sample_name="单样品",
            status=1
        )
        db_session.add(check_object)
        db_session.commit()

        # Add 4 items
        for i in range(4):
            item = CheckObjectItem(
                check_object_id=check_object.id,
                check_item_name=f"项{i}"
            )
            db_session.add(item)
        db_session.commit()

        service = ExcelExportService(db_session)
        count = service.calculate_row_count([check_object.id])

        # Should be 4 rows (4 items)
        assert count == 4

    def test_calculate_row_count_multiple_samples(self, db_session):
        """Test row count calculation for multiple samples"""
        ids = []
        for i in range(5):
            check_object = CheckObject(
                check_no=f"COUNT_MULTI_{i:03d}",
                sample_name=f"样品{i}",
                status=1
            )
            db_session.add(check_object)
            db_session.commit()

            # Each sample has 3 items
            for j in range(3):
                item = CheckObjectItem(
                    check_object_id=check_object.id,
                    check_item_name=f"项{j}"
                )
                db_session.add(item)
            ids.append(check_object.id)

        db_session.commit()

        service = ExcelExportService(db_session)
        count = service.calculate_row_count(ids)

        # Should be 15 rows (5 samples * 3 items)
        assert count == 15

    def test_calculate_row_count_sample_without_items(self, db_session):
        """Test row count for sample without items"""
        check_object = CheckObject(
            check_no="COUNT_NO_ITEMS_001",
            sample_name="无项样品",
            status=1
        )
        db_session.add(check_object)
        db_session.commit()

        service = ExcelExportService(db_session)
        count = service.calculate_row_count([check_object.id])

        # Should be 1 row (empty row for sample)
        assert count == 1

    def test_header_row_generation(self, db_session):
        """Test that header row has correct columns"""
        check_object = CheckObject(
            check_no="HEADER_TEST_001",
            sample_name="表头测试",
            status=1
        )
        db_session.add(check_object)
        db_session.commit()

        item = CheckObjectItem(
            check_object_id=check_object.id,
            check_item_name="项目"
        )
        db_session.add(item)
        db_session.commit()

        service = ExcelExportService(db_session)
        excel_bytes = service.export_to_excel([check_object.id])

        wb = load_workbook(BytesIO(excel_bytes))
        ws = wb.active

        # Verify header row
        expected_headers = [
            "样品名称", "公司/个体", "检测项目", "检验结果",
            "该项结果", "检测时间", "样品编号", "检测方法"
        ]

        for i, header in enumerate(expected_headers, 1):
            cell_value = ws.cell(row=1, column=i).value
            assert cell_value == header, f"Column {i} header mismatch: expected '{header}', got '{cell_value}'"

    def test_workbook_creation(self, db_session):
        """Test that workbook is created correctly"""
        check_object = CheckObject(
            check_no="WB_TEST_001",
            sample_name="工作簿测试",
            status=1
        )
        db_session.add(check_object)
        db_session.commit()

        item = CheckObjectItem(
            check_object_id=check_object.id,
            check_item_name="项目"
        )
        db_session.add(item)
        db_session.commit()

        service = ExcelExportService(db_session)
        excel_bytes = service.export_to_excel([check_object.id])

        # Should be valid Excel file
        assert isinstance(excel_bytes, bytes)
        assert len(excel_bytes) > 0

        # Should be loadable
        wb = load_workbook(BytesIO(excel_bytes))
        assert wb is not None
        assert wb.active is not None

    def test_empty_id_list(self, db_session):
        """Test behavior with empty ID list"""
        service = ExcelExportService(db_session)

        # Should handle gracefully
        count = service.calculate_row_count([])
        assert count == 0

    def test_nonexistent_id(self, db_session):
        """Test behavior with nonexistent ID"""
        service = ExcelExportService(db_session)

        # Should handle gracefully
        count = service.calculate_row_count([99999])
        assert count == 0

    def test_mixed_valid_invalid_ids(self, db_session):
        """Test behavior with mix of valid and invalid IDs"""
        check_object = CheckObject(
            check_no="MIXED_TEST_001",
            sample_name="混合测试",
            status=1
        )
        db_session.add(check_object)
        db_session.commit()

        item = CheckObjectItem(
            check_object_id=check_object.id,
            check_item_name="项目"
        )
        db_session.add(item)
        db_session.commit()

        service = ExcelExportService(db_session)

        # Mix valid and invalid IDs
        ids = [check_object.id, 99998, 99999]
        count = service.calculate_row_count(ids)

        # Should only count valid ID
        assert count == 1
