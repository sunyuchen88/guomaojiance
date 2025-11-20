"""
Excel Export Service
T139-T146: Generate Excel export with openpyxl
"""
from typing import List, Dict, Any, Optional
from datetime import datetime
from io import BytesIO
from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from app.models.check_object import CheckObject
from app.models.check_item import CheckObjectItem


class ExcelExportService:
    """Service for exporting check data to Excel format"""

    # T145: Column headers (8 columns)
    HEADERS = [
        "样品名称",
        "公司/个体",
        "检测项目",
        "检验结果",
        "该项结果",
        "检测时间",
        "样品编号",
        "检测方法"
    ]

    def __init__(self, db: Session):
        self.db = db

    def export_to_excel(
        self,
        check_object_ids: List[int] = None,
        filters: Dict[str, Any] = None
    ) -> bytes:
        """
        T139: Export check data to Excel file

        Args:
            check_object_ids: List of specific IDs to export
            filters: Query filters (status, company, date range)

        Returns:
            Excel file as bytes
        """
        # Get check objects
        check_objects = self._get_check_objects(check_object_ids, filters)

        # Expand to rows
        rows = self.expand_check_items_to_rows(check_objects)

        # T140: Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "检测结果"

        # Write headers
        self._write_headers(ws)

        # Write data rows
        self._write_data_rows(ws, rows)

        # Apply formatting
        self._apply_formatting(ws)

        # Save to bytes
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return output.getvalue()

    def expand_check_items_to_rows(
        self,
        check_objects: List[CheckObject]
    ) -> List[Dict[str, Any]]:
        """
        T139: Expand check items to individual rows
        Each check item becomes one row with sample info repeated

        Args:
            check_objects: List of CheckObject models

        Returns:
            List of row dictionaries
        """
        rows = []

        for obj in check_objects:
            # Get check items for this object
            items = self.db.query(CheckObjectItem).filter(
                CheckObjectItem.check_object_id == obj.id
            ).all()

            if not items:
                # Sample without items - create one row with empty item fields
                rows.append(self._create_row(obj, None))
            else:
                # Create one row per item
                for item in items:
                    rows.append(self._create_row(obj, item))

        return rows

    def _create_row(
        self,
        check_object: CheckObject,
        check_item: Optional[CheckObjectItem]
    ) -> Dict[str, Any]:
        """Create a row dictionary from check object and item"""
        return {
            "sample_name": check_object.sample_name or "",
            "company_name": check_object.company_name or "",
            "check_item_name": check_item.check_item_name if check_item else "",
            "check_result": check_object.check_result or "",
            "item_result": check_item.check_result if check_item else "",
            "sampling_time": self._format_date(check_object.sampling_time),
            "check_no": check_object.check_no or "",
            "check_method": check_item.check_method if check_item else ""
        }

    def _format_date(self, dt: Optional[datetime]) -> str:
        """Format datetime for Excel"""
        if not dt:
            return ""
        return dt.strftime("%Y-%m-%d %H:%M")

    def _get_check_objects(
        self,
        check_object_ids: List[int] = None,
        filters: Dict[str, Any] = None
    ) -> List[CheckObject]:
        """Get check objects by IDs or filters"""
        query = self.db.query(CheckObject)

        if check_object_ids:
            query = query.filter(CheckObject.id.in_(check_object_ids))
        elif filters:
            # T144: Apply query filters
            if filters.get("status") is not None:
                query = query.filter(CheckObject.status == filters["status"])
            if filters.get("company"):
                query = query.filter(
                    CheckObject.company_name.ilike(f"%{filters['company']}%")
                )
            if filters.get("check_no"):
                query = query.filter(CheckObject.check_no == filters["check_no"])
            if filters.get("start_date"):
                query = query.filter(
                    CheckObject.sampling_time >= filters["start_date"]
                )
            if filters.get("end_date"):
                query = query.filter(
                    CheckObject.sampling_time <= filters["end_date"]
                )

        return query.all()

    def calculate_row_count(self, check_object_ids: List[int]) -> int:
        """
        Calculate total row count for limit validation

        Args:
            check_object_ids: List of check object IDs

        Returns:
            Total number of rows that would be generated
        """
        if not check_object_ids:
            return 0

        total = 0

        for obj_id in check_object_ids:
            # Check if object exists
            obj = self.db.query(CheckObject).filter(
                CheckObject.id == obj_id
            ).first()

            if not obj:
                continue

            # Count items
            item_count = self.db.query(CheckObjectItem).filter(
                CheckObjectItem.check_object_id == obj_id
            ).count()

            # At least 1 row per sample
            total += max(1, item_count)

        return total

    def _write_headers(self, ws):
        """Write header row to worksheet"""
        for col, header in enumerate(self.HEADERS, 1):
            cell = ws.cell(row=1, column=col, value=header)
            # Apply header styling
            cell.font = Font(bold=True, size=11)
            cell.fill = PatternFill(
                start_color="4472C4",
                end_color="4472C4",
                fill_type="solid"
            )
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center", vertical="center")

    def _write_data_rows(self, ws, rows: List[Dict[str, Any]]):
        """Write data rows to worksheet"""
        for row_idx, row_data in enumerate(rows, 2):
            ws.cell(row=row_idx, column=1, value=row_data["sample_name"])
            ws.cell(row=row_idx, column=2, value=row_data["company_name"])
            ws.cell(row=row_idx, column=3, value=row_data["check_item_name"])
            ws.cell(row=row_idx, column=4, value=row_data["check_result"])
            ws.cell(row=row_idx, column=5, value=row_data["item_result"])
            ws.cell(row=row_idx, column=6, value=row_data["sampling_time"])
            ws.cell(row=row_idx, column=7, value=row_data["check_no"])
            ws.cell(row=row_idx, column=8, value=row_data["check_method"])

    def _apply_formatting(self, ws):
        """Apply formatting to worksheet"""
        # Set column widths
        column_widths = [20, 25, 20, 12, 15, 18, 20, 20]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + col)].width = width

        # Apply borders to all cells
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=8):
            for cell in row:
                cell.border = thin_border
                if cell.row > 1:
                    cell.alignment = Alignment(
                        horizontal="left",
                        vertical="center",
                        wrap_text=True
                    )


def generate_export_filename() -> str:
    """
    T146: Generate export filename with current date
    Format: 检测结果导出_YYYYMMDD.xlsx
    """
    date_str = datetime.now().strftime("%Y%m%d")
    return f"检测结果导出_{date_str}.xlsx"
